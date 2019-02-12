#!/usr/bin/python3

import argparse
import openpyxl
import pathlib
import itertools
import math
import ast

from openpyxl.styles import Font, Alignment

'''
grader.py grades quiz rubrik

'''

# Should convert this to config file later
cfg = {}
cfg['grades_student_col'] = 1
cfg['grades_skill_row'] = 1
cfg['quiz_student_col'] = 1
cfg['quiz_student_begin'] = 9
cfg['quiz_question_begin'] = 5


def parse_arguments():
	parser = argparse.ArgumentParser(description="Grader for REEF Quizes")
	parser.add_argument(
		'grades',
		help='Excel workbook containing the class grades.'
	)
	parser.add_argument(
		'quiz',
		help='Excel workbook containing the quiz results.'
	)
	parser.add_argument(
		'rubrik',
		help='Grading rubrik in dictionary form.'
	)
	return parser.parse_args()


def find_student(sheet, name):
	'''Find the row which matches the students name.'''
	for i in range(sheet.min_row, sheet.max_row+1):
		cell = sheet.cell(row=i, column=cfg['grades_student_col']).value
		if bool(cell) and name.casefold() == cell.casefold():
			return i


def find_skill_col(sheet, skill):
	'''Find the column which matches the skill name.'''
	for i in range(sheet.min_column, sheet.max_column+1):
		cell = sheet.cell(row=cfg['grades_skill_row'], column=i).value
		if bool(cell) and skill.casefold() == cell.casefold():
			return i


def compute_grade(values):
	'''Computes the averages of the scores and rounds up to the nearest 0.5 value.'''
	average = math.fsum(values) / len(values)
	return round(average / 0.5) * 0.5


def create_new_sheet(wb, name, skills):
	'''Create a new worksheet in the grade workbook and copy over the first 3 columns
	of the first worksheet (i.e. student name, id, and git id).'''
	wb.create_sheet(name)
	new_sheet = wb[name]
	first_sheet = wb[wb.sheetnames[0]]

	columns = list(itertools.islice(first_sheet.columns, 3))
	for i, col in enumerate(columns, 1):
		#new_sheet.column_dimensions[i].width = 17
		for j, cell in enumerate(col, 1):
			if col == 1:
				new_sheet.cell(row=j, column=i).font = Font(bold=True)
			new_sheet.cell(row=j, column=i).value = cell.value

	for c, s in enumerate(skills, 4):
		new_sheet.cell(row=cfg['grades_skill_row'], column=c).value = s
		new_sheet.cell(row=cfg['grades_skill_row'], column=c).alignment = Alignment(text_rotation=45)
		#new.sheet.column_dimensions[c].width = 4

	return new_sheet


def main():
	args = parse_arguments()

	# Check file extensions
	grades = pathlib.Path(args.grades)
	assert (grades.suffix == '.xlsx'), 'Invalid grade workbook!'

	quiz = pathlib.Path(args.quiz)
	assert (quiz.suffix == '.xlsx'), 'Invalid quiz workbook!'

	# Load workbook containing grades
	grades_wb = openpyxl.load_workbook(str(grades))

	# Load workbook containing quiz results
	quiz_wb = openpyxl.load_workbook(str(quiz))
	quiz_ws = quiz_wb['Summary']

	# Read in the grading rubrik
	with open(args.rubrik, 'r') as f:
		rubrik = ast.literal_eval(f.read())
	assert rubrik, 'Could not load rubrik!'

	# Create new worksheet in grade book
	name = quiz.stem + '-quiz'
	grades_ws = create_new_sheet(grades_wb, name, rubrik)

	# Compute and assign grades
	lost = []
	for i in range(cfg['quiz_student_begin'], quiz_ws.max_row):
		student = quiz_ws.cell(row=i, column=cfg['quiz_student_col']).value
		print('Computing grades for {}'.format(student))

		found = find_student(grades_ws, student)
		if found:
			for skill, questions in rubrik.items():
				print('... measuring skill: {}'.format(skill))

				values = []
				for q in questions:
					score = quiz_ws.cell(row=i, column=cfg['quiz_question_begin'] + q - 1)
					if not isinstance(score.value, str):
						print('... using value from {}'.format(score.coordinate))
						values.append(float(score.value))

				if values:
					grade = compute_grade(values)
					print('... computed grade: {}'.format(grade))

					grades_ws.cell(row=found, column=find_skill_col(grades_ws, skill)).value = grade
					print('... grade updated')
		else:
			lost.append(student)
	print('Lost students: {}'.format(lost))
	
	grades_wb.save('test.xlsx')

main()